from openpyxl import Workbook
from openpyxl.styles import Font


def create_scenario_manager_sheet(wb):

    ws = wb.create_sheet("Scenario Manager")

    # ==========================================================
    # TITLE
    # ==========================================================

    ws["A1"] = "Scenario Manager"
    ws["A1"].font = Font(bold=True, size=14)

    # ==========================================================
    # PRODUCTION COST MODIFIERS
    # ==========================================================

    ws["A3"] = "Production Cost Modifiers"

    headers = [
        "Product ID",
        "Distribution",
        "Product Name",
        "Modifier"
    ]

    for col, value in enumerate(headers, start=1):
        ws.cell(4, col, value)

    production_rows = [
        ["BR-1", "retail", "Pilsner", 1.5],
        ["BR-2", "retail", "Bavarian Lager", 1.5],
        ["BR-3", "retail", "Light Wheat", 1.5],
        ["BR-4", "retail", "Red Wheat", 1.5],
        ["BR-5", "retail", "Budweiser", 1.5],
        ["BR-6", "retail", "Samuel Adams", 1.5],
        ["BR-7", "wholesale", "Budweiser", 2.0],
        ["BR-8", "wholesale", "Samuel Adams", 2.0],
    ]

    row = 5

    for item in production_rows:

        ws.cell(row, 1, item[0])
        ws.cell(row, 2, item[1])
        ws.cell(row, 3, item[2])
        ws.cell(row, 4, item[3])

        row += 1

    # ==========================================================
    # ADVERTISING MODIFIERS
    # ==========================================================

    ws["F3"] = "Advertising Cost Modifiers"

    ad_headers = [
        "Name",
        "Modifier"
    ]

    for col, value in enumerate(ad_headers, start=6):
        ws.cell(4, col, value)

    ad_rows = [
        ["Website", 2.0],
        ["Trade Shows", 2.0],
        ["Local Advertising", 2.0],
        ["Firefly", 2.0],
        ["Billboards", 2.0],
        ["Sponsoring Sports Event", 2.0],
        ["Social Media", 2.0],
    ]

    row = 5

    for item in ad_rows:

        ws.cell(row, 6, item[0])
        ws.cell(row, 7, item[1])

        row += 1

    # ==========================================================
    # OPTIMISTIC SCENARIO
    # ==========================================================

    ws["J3"] = "Optimistic Scenario"

    ws["J4"] = "Production Modifier"
    ws["K4"] = 0.75

    ws["J5"] = "Advertising Modifier"
    ws["K5"] = 1.25

    # ==========================================================
    # PESSIMISTIC SCENARIO
    # ==========================================================

    ws["J8"] = "Pessimistic Scenario"

    ws["J9"] = "Production Modifier"
    ws["K9"] = 1.50

    ws["J10"] = "Advertising Modifier"
    ws["K10"] = 0.75

    # ==========================================================
    # NOTES
    # ==========================================================

    ws["A16"] = (
        "Excel Scenario Manager originally changed "
        "AG17:AG24 and AG29:AG35."
    )

    ws["A17"] = (
        "These rows represent the production "
        "and advertising scenario modifiers."
    )

    return ws