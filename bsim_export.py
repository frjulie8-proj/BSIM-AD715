"""
bsim_export.py
--------------
Generates a BSIM Scenario Analysis Excel file from webapp export data.

Usage:
    export_bsim_excel(
        fixed_cost=441002.73,
        var_cost=0.92,
        avg_price=3.62,
        qty_actual=405480,
        output_path="BSIM_Output.xlsx"
    )
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.formatting.rule import ColorScaleRule

# ── Shared styles ─────────────────────────────────────────────────────────────
FONT_NAME       = "Arial"
NAVY            = "2E4057"
BLUE            = "4A90D9"
YELLOW          = "FFFF00"
LIGHT_BLUE      = "EBF3FB"
GRAY            = "F5F5F5"
GREEN_BG        = "D4EFDF"
RED_BG          = "FADBD8"
COL_WIDTH_LABEL = 28   # Fix 2: shared constant — both sheets use this
COL_WIDTH_VALUE = 16   # Fix 2: shared constant — both sheets use this

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(top=s, bottom=s, left=s, right=s)

def _font(bold=False, color="000000", size=11, italic=False):
    return Font(name=FONT_NAME, bold=bold, color=color, size=size, italic=italic)

def _fill(color):
    return PatternFill("solid", start_color=color)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _style(cell, value, font, fill, align, border, fmt=None):
    cell.value = value
    cell.font = font
    cell.fill = fill
    cell.alignment = align
    cell.border = border
    if fmt:
        cell.number_format = fmt


# ── Function 0a: write_header ─────────────────────────────────────────────────
def write_header(ws, start_row, title, subtitle):
    """
    Writes a 2-row header block merged across A:C.
    Returns the next available row after 1 blank row gap.
    """
    border = _thin_border()

    ws.merge_cells(f"A{start_row}:C{start_row}")
    _style(ws.cell(start_row, 1), title,
           _font(bold=True, color="FFFFFF", size=12),
           _fill(NAVY), _align(wrap=True), border)
    ws.row_dimensions[start_row].height = 28

    sub_row = start_row + 1
    ws.merge_cells(f"A{sub_row}:C{sub_row}")
    _style(ws.cell(sub_row, 1), subtitle,
           _font(bold=True, color="FFFFFF", size=10),
           _fill(BLUE), _align(wrap=True), border)

    return sub_row + 2  # +1 blank row gap


# ── Function 0b: write_instructions ──────────────────────────────────────────
def write_instructions(ws, start_row, title, steps):
    """
    Writes a titled instruction block: title row + one row per step.
    Returns the next available row after 1 blank row gap.
    """
    border = _thin_border()

    ws.merge_cells(f"A{start_row}:C{start_row}")
    _style(ws.cell(start_row, 1), title,
           _font(bold=True, color="FFFFFF", size=11),
           _fill(NAVY), _align(), border)
    ws.row_dimensions[start_row].height = 24

    for i, (label, desc) in enumerate(steps):
        r = start_row + 1 + i
        _style(ws.cell(r, 1), label,
               _font(bold=True, color=NAVY, size=10),
               _fill(GRAY), _align(), border)
        ws.merge_cells(f"B{r}:C{r}")
        _style(ws.cell(r, 2), desc,
               _font(color="333333", size=10),
               _fill("FFFFFF"), _align(h="left"), border)

    last_row = start_row + len(steps)
    return last_row + 2  # +1 blank row gap


# ── Function 0c: write_inputs ─────────────────────────────────────────────────
def write_inputs(ws, start_row, inputs):
    """
    Writes a labelled INPUTS section.
    inputs: list of (key, label, value, fmt, note)
    Returns (next_row, dict) — dict keyed by 'key' for safe formula references.
    Reorder-safe: the key travels with its data row, no positional assumptions.
    """
    border  = _thin_border()
    blue_in = _font(color="0000FF", size=11)
    navy_b  = _font(bold=True, color=NAVY, size=11)
    gray_sm = _font(color="888888", size=9, italic=True)

    # Section label
    ws.merge_cells(f"A{start_row}:C{start_row}")
    _style(ws.cell(start_row, 1), "INPUTS  (blue = hardcoded, change to test scenarios)",
           _font(bold=True, color="FFFFFF", size=11),
           _fill(BLUE), _align(), border)
    row = start_row + 1

    input_rows = {}
    for key, label, val, fmt, note in inputs:
        _style(ws.cell(row, 1), label, navy_b,  _fill(GRAY),   _align(h="left"), border)
        _style(ws.cell(row, 2), val,   blue_in, _fill(YELLOW), _align(),         border, fmt)
        if note:
            ws.cell(row, 3).value     = note
            ws.cell(row, 3).font      = gray_sm
            ws.cell(row, 3).alignment = _align(h="left")
        input_rows[key] = row
        row += 1

    return row + 1, input_rows  # +1 blank row gap


# ── Function 0d: write_recommendation ─────────────────────────────────────────
def write_recommendation(ws, start_row, anchor_col="E",
                         title="Business Recommendation", n_body_rows=7, width=3):
    """
    Writes a reusable, empty Business Recommendation block to the right of the
    sheet's data. DRY: same call works on any sheet, only the anchor changes.
    No content is baked in — the body rows are blank for the user to fill in.

    anchor_col  : left-most column of the block as a letter (e.g. "E", "J")
    n_body_rows : number of empty wrap-text rows under the title
    width       : number of columns the block spans
    Returns next available row after a 1-row gap.
    """
    no_border = Border()  # borderless — clean merged block, no interior lines
    c0 = column_index_from_string(anchor_col)
    c1 = c0 + width - 1
    left  = get_column_letter(c0)
    right = get_column_letter(c1)

    # Widen spanned columns so notes are readable
    for c in range(c0, c1 + 1):
        col = get_column_letter(c)
        if (ws.column_dimensions[col].width or 0) < 18:
            ws.column_dimensions[col].width = 18

    # Title bar — single merged row, no border
    for c in range(c0, c1 + 1):
        _style(ws.cell(start_row, c), title if c == c0 else None,
               _font(bold=True, color="FFFFFF", size=12), _fill(NAVY),
               _align(wrap=True), no_border)
    ws.merge_cells(f"{left}{start_row}:{right}{start_row}")
    ws.row_dimensions[start_row].height = 28

    # Body — ONE merged block (no interior lines), light fill, free-form notes
    body_top = start_row + 1
    body_bot = start_row + n_body_rows
    for r in range(body_top, body_bot + 1):
        for c in range(c0, c1 + 1):
            _style(ws.cell(r, c), None,
                   _font(color="333333", size=10), _fill(LIGHT_BLUE),
                   _align(h="left", v="top", wrap=True), no_border)
    ws.merge_cells(f"{left}{body_top}:{right}{body_bot}")

    return body_bot + 2  # +1 blank row gap


# ── Function 1: generate_series ───────────────────────────────────────────────
def generate_series(midpoint, n_steps, pct=0.10):
    """
    Returns a list of n_steps*2+1 values centered around midpoint.
    Increment is computed as pct * midpoint (percentage-based, never flat).

    Example: midpoint=3.62, n_steps=3, pct=0.10
             increment = 0.362
             → [2.534, 2.896, 3.258, 3.62, 3.982, 4.344, 4.706]
    """
    increment = midpoint * pct
    values = []
    for i in range(-n_steps, n_steps + 1):
        values.append(round(midpoint + i * increment, 4))
    return values


# ── Function 1b: generate_breakeven_series ────────────────────────────────────
def generate_breakeven_series(qty_actual, n_points=10, frac=1.2):
    """
    Returns an ascending quantity axis from 0 up to ~frac*qty_actual,
    in n_points even steps (n_points+1 values, starting at 0).
    Scales with the BSIM scenario — never a flat/hardcoded range.
    """
    qty_max = round(frac * qty_actual)
    step = qty_max / n_points
    return [round(i * step) for i in range(n_points + 1)]


# ── Function 1c: generate_breakeven_chart_data ────────────────────────────────
def generate_breakeven_chart_data(ws, start_row, input_rows, qty_actual,
                                  n_points=10, frac=1.2, chart_anchor="E"):
    """
    Writes the Break Even chart-data table (Quantity | Revenue | Cost) with
    fully dynamic formulas referencing the shared INPUTS, then draws a line chart.
    Revenue = avg_price * qty ; Cost = fixed_cost + var_cost * qty.
    input_rows: dict from write_inputs (keys: fixed, var, avg) — reorder-safe.
    Returns next available row after a blank gap.
    """
    border = _thin_border()
    navy_b = _font(bold=True, color=NAVY, size=11)
    black_f = _font(color="000000", size=11)

    r_fixed = input_rows["fixed"]
    r_var   = input_rows["var"]
    r_avg   = input_rows["avg"]

    # Section label
    ws.merge_cells(f"A{start_row}:C{start_row}")
    _style(ws.cell(start_row, 1), "CHART DATA — Break Even Line Chart",
           _font(bold=True, color="FFFFFF", size=11),
           _fill(BLUE), _align(), border)

    # Column headers
    header_row = start_row + 1
    for col, label in enumerate(
            ["Quantity (pints)", "Total Revenue ($)", "Total Cost ($)"], start=1):
        _style(ws.cell(header_row, col), label,
               _font(bold=True, color="FFFFFF", size=11),
               _fill(BLUE), _align(), border)

    # Data rows — qty is a hardcoded axis; revenue/cost are dynamic formulas
    qty_series = generate_breakeven_series(qty_actual, n_points, frac)
    first_data_row = header_row + 1
    r = first_data_row
    for qty in qty_series:
        _style(ws.cell(r, 1), qty,                       black_f, _fill(GRAY),       _align(), border, "#,##0")
        _style(ws.cell(r, 2), f"=B{r_avg}*A{r}",         black_f, _fill(LIGHT_BLUE), _align(), border, "$#,##0")
        _style(ws.cell(r, 3), f"=B{r_fixed}+B{r_var}*A{r}", black_f, _fill(LIGHT_BLUE), _align(), border, "$#,##0")
        r += 1
    last_data_row = r - 1

    # Line chart: Revenue & Cost vs Quantity
    chart = LineChart()
    chart.title = "Break Even Analysis"
    chart.style = 2
    chart.x_axis.title = "Quantity (pints)"
    chart.y_axis.title = "USD"
    # Faint gridlines back — B3B3B3 ≈ 30% opacity black on white (openpyxl can't
    # set a true alpha transform, so this is the render-safe visual equivalent)
    def _faint_grid():
        ln = LineProperties(solidFill="B3B3B3", w=9525)  # 0.75pt thin line
        return ChartLines(spPr=GraphicalProperties(ln=ln))
    chart.x_axis.majorGridlines = _faint_grid()
    chart.y_axis.majorGridlines = _faint_grid()
    # Show numeric axis values (openpyxl hides them unless axes are forced on)
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "nextTo"
    chart.y_axis.numFmt = "$#,##0"
    chart.height = 11   # larger so title + axis values are legible, not overlapping
    chart.width  = 21
    data = Reference(ws, min_col=2, max_col=3, min_row=header_row, max_row=last_data_row)
    cats = Reference(ws, min_col=1, min_row=first_data_row, max_row=last_data_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"{chart_anchor}{start_row}")

    return last_data_row + 2  # +1 blank row gap


# ── Function 2: generate_sensitivity_table ────────────────────────────────────
def generate_sensitivity_table(ws, start_row,
                                corner_formula,
                                col_series, row_series,
                                col_fmt, row_fmt,
                                instr_title, instr_steps):
    """
    Generic sensitivity table writer.
    Writes: corner formula cell | col_series headers | row_series headers | blank data cells.
    Followed by instruction block.

    Args:
        corner_formula : Excel formula string for the top-left corner cell
        col_series     : list of values for column headers (row of corner)
        row_series     : list of values for row headers (column of corner)
        col_fmt        : number format for column headers
        row_fmt        : number format for row headers
        instr_title    : title string for HOW TO RUN block
        instr_steps    : list of (label, desc) tuples

    Returns next available row after blank gap.
    """
    border  = _thin_border()
    black_f = _font(color="000000", size=11)
    white_b = _font(bold=True, color="FFFFFF", size=11)

    r_corner = start_row

    # Corner formula cell
    _style(ws.cell(r_corner, 1), corner_formula,
           black_f, _fill(LIGHT_BLUE), _align(), border, "$#,##0.00")

    # Column headers (same row as corner, col 2 onwards)
    for j, val in enumerate(col_series, start=2):
        _style(ws.cell(r_corner, j), val,
               white_b, _fill(BLUE), _align(), border, col_fmt)

    row = r_corner + 1

    # Row headers + blank data cells
    for rv in row_series:
        _style(ws.cell(row, 1), rv,
               white_b, _fill(BLUE), _align(), border, row_fmt)
        for j in range(2, len(col_series) + 2):
            ws.cell(row, j).fill          = _fill(LIGHT_BLUE)
            ws.cell(row, j).alignment     = _align()
            ws.cell(row, j).border        = border
            ws.cell(row, j).number_format = "$#,##0"
        row += 1

    # Heatmap — red (loss) → yellow → green (high profit) across the data grid
    first_data_row = r_corner + 1
    last_data_row  = r_corner + len(row_series)
    last_data_col  = get_column_letter(len(col_series) + 1)
    data_range = f"B{first_data_row}:{last_data_col}{last_data_row}"
    ws.conditional_formatting.add(
        data_range,
        ColorScaleRule(
            start_type="min",        start_color=RED_BG,
            mid_type="percentile",   mid_value=50, mid_color="FFFFCC",
            end_type="max",          end_color=GREEN_BG,
        )
    )

    row += 1  # blank row

    # Instructions
    row = write_instructions(ws, row, instr_title, instr_steps)

    return row


# ── Function 3: generate_goal_seek ────────────────────────────────────────────
def generate_goal_seek(ws, start_row, fixed_cost, var_cost, avg_price, qty_actual):
    """
    Writes the Goal Seek - Break Even sheet.
    Returns next available row after blank gap.
    """
    border  = _thin_border()
    black_f = _font(color="000000", size=11)
    navy_b  = _font(bold=True, color=NAVY, size=11)
    gray_sm = _font(color="888888", size=9, italic=True)

    # Column widths — Fix 2: use shared constants
    ws.column_dimensions["A"].width = COL_WIDTH_LABEL
    ws.column_dimensions["B"].width = COL_WIDTH_VALUE
    ws.column_dimensions["C"].width = 30

    # Header
    row = write_header(
        ws, start_row,
        "BSIM Scenario Analysis — Goal Seek: Break Even",
        "Source: 36 Month Pro Forma — FY1 Data  |  AD715 BSIM Project"
    )

    # Inputs — Fix 1: key-based tuples, dict return
    row, ir = write_inputs(ws, row, [
        ("fixed", "Total Fixed Cost FY1 ($)",     fixed_cost, "$#,##0.00", None),
        ("var",   "Variable Cost per Pint ($)",   var_cost,   "$#,##0.00", None),
        ("avg",   "Average Revenue per Pint ($)", avg_price,  "$#,##0.00", None),
    ])

    # Section label: CALCULATIONS
    ws.merge_cells(f"A{row}:C{row}")
    _style(ws.cell(row, 1), "CALCULATIONS  (black = formula)",
           _font(bold=True, color="FFFFFF", size=11),
           _fill(BLUE), _align(), border)
    row += 1

    # Derive row refs dynamically
    r_qty  = row
    r_rev  = row + 1
    r_vc   = row + 2
    r_cont = row + 3
    r_fc   = row + 4
    r_gp   = row + 5
    r_tax  = row + 6
    r_np   = row + 7

    calcs = [
        ("Total Quantity Sold (pints)", qty_actual,                                 "#,##0",     GREEN_BG,   "← Change this cell in Goal Seek (green)"),
        ("Total Revenue ($)",           f"=B{ir['avg']}*B{r_qty}",            "$#,##0.00", LIGHT_BLUE, "Avg Price x Quantity"),
        ("Total Variable Cost ($)",     f"=B{ir['var']}*B{r_qty}",            "$#,##0.00", LIGHT_BLUE, "Var Cost x Quantity"),
        ("Total Contribution ($)",      f"=B{r_rev}-B{r_vc}",                 "$#,##0.00", LIGHT_BLUE, "Revenue minus Variable Cost"),
        ("Total Fixed Cost ($)",        f"=B{ir['fixed']}",                   "$#,##0.00", LIGHT_BLUE, "From inputs above"),
        ("Gross Profit ($)",            f"=B{r_cont}-B{r_fc}",                "$#,##0.00", RED_BG,     "← Set this to 0 in Goal Seek (red)"),
        ("Tax (21%)",                   f"=IF(B{r_gp}>0,B{r_gp}*0.21,0)",   "$#,##0.00", LIGHT_BLUE, ""),
        ("Net Profit ($)",              f"=B{r_gp}-B{r_tax}",                 "$#,##0.00", LIGHT_BLUE, ""),
    ]

    for label, val, fmt, bg, note in calcs:
        _style(ws.cell(row, 1), label, navy_b,  _fill(GRAY), _align(h="left"), border)
        _style(ws.cell(row, 2), val,   black_f, _fill(bg),   _align(),         border, fmt)
        if note:
            ws.cell(row, 3).value     = note
            ws.cell(row, 3).font      = gray_sm
            ws.cell(row, 3).alignment = _align(h="left")
        row += 1

    row += 1  # blank row

    steps = [
        ("Step 1", f"Click on cell B{r_gp} (Gross Profit — highlighted red)"),
        ("Step 2", "Go to Data tab → What-If Analysis → Goal Seek"),
        ("Step 3", f"Set cell: B{r_gp}  |  To value: 0  |  By changing cell: B{r_qty}"),
        ("Step 4", "Click OK — Excel will find the break even quantity"),
        ("Step 5", "Note the break even quantity and compare to your actual FY1 quantity"),
    ]
    row = write_instructions(ws, row, "HOW TO RUN GOAL SEEK", steps)

    return row, ir


# ── Function 4a: generate_table_price_qty ────────────────────────────────────
def generate_table_price_qty(ws, start_row, input_rows, avg_price, qty_actual,
                              n_steps=3, pct=0.10):
    """
    Table 1: Price x Quantity.
    Rows = price series, Columns = qty series.
    Variable cost and fixed cost held at actual values from shared inputs.
    input_rows: dict returned by write_inputs — keyed by name, reorder-safe.
    """
    r_fixed = input_rows["fixed"]
    r_var   = input_rows["var"]
    r_avg   = input_rows["avg"]
    r_qty   = input_rows["qty"]

    qty_series   = generate_series(qty_actual, n_steps, pct)
    price_series = generate_series(avg_price,  n_steps, pct)

    corner_formula = f"=(B{r_avg}-B{r_var})*B{r_qty}-B{r_fixed}"
    last_data_col  = get_column_letter(len(qty_series) + 1)
    last_data_row  = start_row + len(price_series)

    steps = [
        ("Step 1", f"Select table range A{start_row}:{last_data_col}{last_data_row}"),
        ("Step 2", "Go to Data tab → What-If Analysis → Data Table"),
        ("Step 3", f"Row input cell: B{r_avg} (Average Revenue per Pint)"),
        ("Step 4", f"Column input cell: B{r_qty} (Total Quantity Sold)"),
        ("Step 5", "Click OK — Excel fills all blank cells automatically"),
        ("Step 6", "Find the row matching your price and read across to find break even quantity"),
    ]

    return generate_sensitivity_table(
        ws, start_row=start_row,
        corner_formula=corner_formula,
        col_series=qty_series,
        row_series=price_series,
        col_fmt="#,##0",
        row_fmt="$#,##0.00",
        instr_title="HOW TO RUN DATA TABLE — Table 1: Price x Quantity",
        instr_steps=steps
    )


# ── Function 4b: generate_table_cost_qty ─────────────────────────────────────
def generate_table_cost_qty(ws, start_row, input_rows, var_cost, qty_actual,
                             n_steps=3, pct=0.10):
    """
    Table 2: Cost x Quantity.
    Rows = var_cost series, Columns = qty series.
    Price and fixed cost held at actual values from shared inputs.
    input_rows: dict returned by write_inputs — keyed by name, reorder-safe.
    """
    r_fixed = input_rows["fixed"]
    r_var   = input_rows["var"]
    r_avg   = input_rows["avg"]
    r_qty   = input_rows["qty"]

    cost_series = generate_series(var_cost,   n_steps, pct)
    qty_series  = generate_series(qty_actual, n_steps, pct)

    corner_formula = f"=(B{r_avg}-B{r_var})*B{r_qty}-B{r_fixed}"
    last_data_col  = get_column_letter(len(qty_series) + 1)
    last_data_row  = start_row + len(cost_series)

    steps = [
        ("Step 1", f"Select table range A{start_row}:{last_data_col}{last_data_row}"),
        ("Step 2", "Go to Data tab → What-If Analysis → Data Table"),
        ("Step 3", f"Row input cell: B{r_var} (Variable Cost per Pint — varies down rows)"),
        ("Step 4", f"Column input cell: B{r_qty} (Total Quantity Sold — varies across columns)"),
        ("Step 5", "Click OK — Excel fills all blank cells automatically"),
        ("Step 6", "Find the row matching your cost and read across to see how quantity affects profit"),
    ]

    return generate_sensitivity_table(
        ws, start_row=start_row,
        corner_formula=corner_formula,
        col_series=qty_series,
        row_series=cost_series,
        col_fmt="#,##0",
        row_fmt="$#,##0.00",
        instr_title="HOW TO RUN DATA TABLE — Table 2: Cost x Quantity",
        instr_steps=steps
    )


# ── Function 4c: generate_table_price_cost ───────────────────────────────────
def generate_table_price_cost(ws, start_row, input_rows, avg_price, var_cost,
                               n_steps=3, pct=0.10):
    """
    Table 3: Price x Cost.
    Rows = var_cost series, Columns = price series.
    Quantity and fixed cost held at actual values from shared inputs.
    input_rows: dict returned by write_inputs — keyed by name, reorder-safe.
    """
    r_fixed = input_rows["fixed"]
    r_var   = input_rows["var"]
    r_avg   = input_rows["avg"]
    r_qty   = input_rows["qty"]

    price_series = generate_series(avg_price, n_steps, pct)
    cost_series  = generate_series(var_cost,  n_steps, pct)

    corner_formula = f"=(B{r_avg}-B{r_var})*B{r_qty}-B{r_fixed}"
    last_data_col  = get_column_letter(len(price_series) + 1)
    last_data_row  = start_row + len(cost_series)

    steps = [
        ("Step 1", f"Select table range A{start_row}:{last_data_col}{last_data_row}"),
        ("Step 2", "Go to Data tab → What-If Analysis → Data Table"),
        ("Step 3", f"Row input cell: B{r_var} (Variable Cost per Pint)"),
        ("Step 4", f"Column input cell: B{r_avg} (Average Revenue per Pint)"),
        ("Step 5", "Click OK — Excel fills all blank cells automatically"),
        ("Step 6", "Find the intersection where margin is healthiest given your cost structure"),
    ]

    return generate_sensitivity_table(
        ws, start_row=start_row,
        corner_formula=corner_formula,
        col_series=price_series,
        row_series=cost_series,
        col_fmt="$#,##0.00",
        row_fmt="$#,##0.00",
        instr_title="HOW TO RUN DATA TABLE — Table 3: Price x Cost",
        instr_steps=steps
    )


# ── Function 4: generate_data_table ──────────────────────────────────────────
def generate_data_table(ws, start_row, fixed_cost, var_cost, avg_price, qty_actual,
                        n_steps=3, pct=0.10):
    """
    Writes the full Data Table - Sensitivity sheet.
    Contains Table 1 (Price x Qty), Table 2 (Cost x Qty), Table 3 (Price x Cost)
    stacked vertically. All series use percentage-based increments.
    Returns next available row after blank gap.
    """
    # Column widths — Fix 2: use shared constants
    ws.column_dimensions["A"].width = COL_WIDTH_LABEL
    n_cols = n_steps * 2 + 2
    for col in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = COL_WIDTH_VALUE

    # Sheet header
    row = write_header(
        ws, start_row,
        "BSIM Sensitivity Analysis — Data Table: Gross Profit by Price and Quantity",
        "Source: 36 Month Pro Forma FY1  |  AD715 BSIM Project"
    )

    # Inputs — shared across all three tables; Fix 1: key-based tuples, dict return
    row, input_rows = write_inputs(ws, row, [
        ("fixed", "Total Fixed Cost FY1 ($)",     fixed_cost, "$#,##0.00", "<- Fixed Cost input"),
        ("var",   "Variable Cost per Pint ($)",   var_cost,   "$#,##0.00", "<- Varies in Cost tables"),
        ("avg",   "Average Revenue per Pint ($)", avg_price,  "$#,##0.00", "<- Varies in Price tables"),
        ("qty",   "Total Quantity Sold (pints)",  qty_actual, "#,##0",     "<- Mid value for Qty series"),
    ])

    # ── Table 1: Price x Quantity ─────────────────────────────────────────────
    row = generate_table_price_qty(
        ws, row,
        input_rows=input_rows,
        avg_price=avg_price, qty_actual=qty_actual,
        n_steps=n_steps, pct=pct
    )

    # ── Table 2: Cost x Quantity ──────────────────────────────────────────────
    row = generate_table_cost_qty(
        ws, row,
        input_rows=input_rows,
        var_cost=var_cost, qty_actual=qty_actual,
        n_steps=n_steps, pct=pct
    )

    # ── Table 3: Price x Cost ─────────────────────────────────────────────────
    row = generate_table_price_cost(
        ws, row,
        input_rows=input_rows,
        avg_price=avg_price, var_cost=var_cost,
        n_steps=n_steps, pct=pct
    )

    return row


# ── Function 5: export_bsim_excel ────────────────────────────────────────────
def export_bsim_excel(fixed_cost, var_cost, avg_price, qty_actual, output_path,
                      n_steps=3, pct=0.10):
    """
    Main entry point. Creates the full BSIM Excel export file.

    Parameters:
        fixed_cost  : Total Fixed Cost FY1 from BSIM webapp
        var_cost    : Variable Cost per Pint from BSIM webapp
        avg_price   : Average Revenue per Pint from BSIM webapp
        qty_actual  : Total Quantity Sold FY1 from BSIM webapp
        output_path : File path for the output .xlsx file
        n_steps     : Number of steps above and below midpoint (default 3 = 7 values)
        pct         : Percentage increment per step (default 10%)
    """
    wb = Workbook()

    # Sheet 1: Goal Seek
    ws1 = wb.active
    ws1.title = "Goal Seek - Break Even"
    row1, ir1 = generate_goal_seek(ws1, start_row=1,
                                   fixed_cost=fixed_cost,
                                   var_cost=var_cost,
                                   avg_price=avg_price,
                                   qty_actual=qty_actual)

    # Chart-data table + Break Even line chart (dynamic formulas)
    generate_breakeven_chart_data(ws1, row1, input_rows=ir1,
                                  qty_actual=qty_actual,
                                  n_points=n_steps * 2 + 4,
                                  chart_anchor="E")

    # Business Recommendation block (DRY) — top-right of sheet 1
    write_recommendation(ws1, start_row=1, anchor_col="E")

    # Sheet 2: Data Table (all sensitivity tables stacked)
    ws2 = wb.create_sheet("Data Table - Sensitivity")
    generate_data_table(ws2, start_row=1,
                        fixed_cost=fixed_cost,
                        var_cost=var_cost,
                        avg_price=avg_price,
                        qty_actual=qty_actual,
                        n_steps=n_steps,
                        pct=pct)

    # Business Recommendation block (DRY) — same position as sheet 1 (E→G, top)
    write_recommendation(ws2, start_row=1, anchor_col="E")

    wb.save(output_path)
    print(f"Saved: {output_path}")

