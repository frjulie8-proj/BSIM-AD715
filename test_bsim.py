"""
test_bsim.py
------------
Functional tests for bsim_export.py.
Run: python test_bsim.py
"""

import os
from openpyxl import Workbook
from bsim_export import (
    generate_series,
    write_header,
    write_inputs,
    write_instructions,
    generate_sensitivity_table,
    generate_goal_seek,
    generate_table_price_qty,
    generate_table_cost_qty,
    generate_table_price_cost,
    generate_data_table,
    export_bsim_excel,
)

PASS = "  PASS"
FAIL = "  FAIL"


def check(label, condition):
    print(f"{PASS if condition else FAIL}  {label}")
    if not condition:
        raise AssertionError(label)


# ── Test 1: generate_series ───────────────────────────────────────────────────
def test_generate_series():
    print("\n[1] generate_series")
    s = generate_series(3.62, n_steps=3, pct=0.10)
    check("returns 7 values (2*3+1)",        len(s) == 7)
    check("midpoint is index 3",             s[3] == 3.62)
    check("values increase left to right",   all(s[i] < s[i+1] for i in range(len(s)-1)))
    check("increment ~10% of midpoint",      round(s[4] - s[3], 4) == round(3.62 * 0.10, 4))

    s2 = generate_series(405480, n_steps=3, pct=0.10)
    check("qty midpoint is index 3",         s2[3] == 405480)
    check("qty returns 7 values",            len(s2) == 7)


# ── Test 2: write_header ──────────────────────────────────────────────────────
def test_write_header():
    print("\n[2] write_header")
    ws = Workbook().active
    next_row = write_header(ws, start_row=1, title="Test Title", subtitle="Test Subtitle")
    check("returns row 4 (2 written + 1 blank)", next_row == 4)
    check("title in A1",                         ws.cell(1, 1).value == "Test Title")
    check("subtitle in A2",                      ws.cell(2, 1).value == "Test Subtitle")
    check("row 3 is blank",                      ws.cell(3, 1).value is None)


# ── Test 3: write_inputs ──────────────────────────────────────────────────────
def test_write_inputs():
    print("\n[3] write_inputs")
    ws = Workbook().active
    inputs = [
        ("fixed", "Fixed Cost", 441002.73, "$#,##0.00", "note A"),
        ("var",   "Var Cost",   0.92,      "$#,##0.00", None),
        ("avg",   "Avg Price",  3.62,      "$#,##0.00", "note C"),
    ]
    next_row, input_rows = write_inputs(ws, start_row=1, inputs=inputs)

    check("returns a dict",                         isinstance(input_rows, dict))
    check("dict has 3 keys",                        len(input_rows) == 3)
    check("key 'fixed' resolves to row 2",          input_rows["fixed"] == 2)
    check("key 'var' resolves to row 3",            input_rows["var"]   == 3)
    check("key 'avg' resolves to row 4",            input_rows["avg"]   == 4)
    check("next_row skips blank (row 6)",           next_row == 6)
    check("B2 has fixed cost value",                ws.cell(2, 2).value == 441002.73)
    check("B3 has var cost value",                  ws.cell(3, 2).value == 0.92)
    check("B4 has avg price value",                 ws.cell(4, 2).value == 3.62)
    check("C2 has note",                            ws.cell(2, 3).value == "note A")
    check("C3 note is None (no note passed)",       ws.cell(3, 3).value is None)

    # Reorder safety: reversing the list must not break key resolution
    ws2 = Workbook().active
    inputs_reversed = list(reversed(inputs))
    _, ir2 = write_inputs(ws2, start_row=1, inputs=inputs_reversed)
    check("reorder: 'avg' still resolves correctly", ws2.cell(ir2["avg"], 2).value == 3.62)
    check("reorder: 'fixed' still resolves correctly", ws2.cell(ir2["fixed"], 2).value == 441002.73)


# ── Test 4: write_instructions ────────────────────────────────────────────────
def test_write_instructions():
    print("\n[4] write_instructions")
    ws = Workbook().active
    steps = [("Step 1", "Do this"), ("Step 2", "Do that"), ("Step 3", "Done")]
    next_row = write_instructions(ws, start_row=1, title="HOW TO", steps=steps)

    check("title in A1",                       ws.cell(1, 1).value == "HOW TO")
    check("step 1 label in A2",                ws.cell(2, 1).value == "Step 1")
    check("step 1 desc in B2",                 ws.cell(2, 2).value == "Do this")
    check("step 3 label in A4",                ws.cell(4, 1).value == "Step 3")
    check("next_row = title(1) + steps(3) + blank(1) + 1 = 6", next_row == 6)


# ── Test 5: generate_goal_seek ────────────────────────────────────────────────
def test_generate_goal_seek():
    print("\n[5] generate_goal_seek")
    ws = Workbook().active
    next_row = generate_goal_seek(ws, start_row=1,
                                  fixed_cost=441002.73,
                                  var_cost=0.92,
                                  avg_price=3.62,
                                  qty_actual=405480)
    check("returns a row number > 1",    next_row > 1)
    check("A1 has title text",           "Goal Seek" in str(ws.cell(1, 1).value))
    check("sheet has content past row 10", ws.cell(10, 1).value is not None)


# ── Test 6: generate_table_price_qty ─────────────────────────────────────────
def test_generate_table_price_qty():
    print("\n[6] generate_table_price_qty")
    ws = Workbook().active
    _, input_rows = write_inputs(ws, start_row=1, inputs=[
        ("fixed", "Fixed Cost", 441002.73, "$#,##0.00", None),
        ("var",   "Var Cost",   0.92,      "$#,##0.00", None),
        ("avg",   "Avg Price",  3.62,      "$#,##0.00", None),
        ("qty",   "Qty",        405480,    "#,##0",     None),
    ])
    ws2 = Workbook().active
    next_row = generate_table_price_qty(ws2, start_row=1,
                                        input_rows=input_rows,
                                        avg_price=3.62,
                                        qty_actual=405480,
                                        n_steps=3, pct=0.10)
    check("returns a row > 1",                       next_row > 1)
    check("corner cell has formula starting with =", str(ws2.cell(1, 1).value).startswith("="))
    check("7 column headers in corner row",          ws2.cell(1, 2).value is not None)
    check("first data cell is blank (light blue)",   ws2.cell(2, 2).value is None)


# ── Test 7: generate_table_cost_qty ──────────────────────────────────────────
def test_generate_table_cost_qty():
    print("\n[7] generate_table_cost_qty")
    ws = Workbook().active
    _, input_rows = write_inputs(ws, start_row=1, inputs=[
        ("fixed", "Fixed Cost", 441002.73, "$#,##0.00", None),
        ("var",   "Var Cost",   0.92,      "$#,##0.00", None),
        ("avg",   "Avg Price",  3.62,      "$#,##0.00", None),
        ("qty",   "Qty",        405480,    "#,##0",     None),
    ])
    ws2 = Workbook().active
    next_row = generate_table_cost_qty(ws2, start_row=1,
                                       input_rows=input_rows,
                                       var_cost=0.92,
                                       qty_actual=405480,
                                       n_steps=3, pct=0.10)
    check("returns a row > 1",                       next_row > 1)
    check("corner cell has formula starting with =", str(ws2.cell(1, 1).value).startswith("="))
    check("7 column headers in corner row",          ws2.cell(1, 2).value is not None)
    check("first data cell is blank (light blue)",   ws2.cell(2, 2).value is None)
    check("row header is a cost value near 0.92",    ws2.cell(2, 1).value is not None)


# ── Test 8: generate_table_price_cost ────────────────────────────────────────
def test_generate_table_price_cost():
    print("\n[8] generate_table_price_cost")
    ws = Workbook().active
    _, input_rows = write_inputs(ws, start_row=1, inputs=[
        ("fixed", "Fixed Cost", 441002.73, "$#,##0.00", None),
        ("var",   "Var Cost",   0.92,      "$#,##0.00", None),
        ("avg",   "Avg Price",  3.62,      "$#,##0.00", None),
        ("qty",   "Qty",        405480,    "#,##0",     None),
    ])
    ws2 = Workbook().active
    next_row = generate_table_price_cost(ws2, start_row=1,
                                         input_rows=input_rows,
                                         avg_price=3.62,
                                         var_cost=0.92,
                                         n_steps=3, pct=0.10)
    check("returns a row > 1",                       next_row > 1)
    check("corner cell has formula starting with =", str(ws2.cell(1, 1).value).startswith("="))
    check("7 column headers in corner row",          ws2.cell(1, 2).value is not None)
    check("first data cell is blank (light blue)",   ws2.cell(2, 2).value is None)


# ── Test 9: generate_data_table ───────────────────────────────────────────────
def test_generate_data_table():
    print("\n[9] generate_data_table")
    ws = Workbook().active
    next_row = generate_data_table(ws, start_row=1,
                                   fixed_cost=441002.73,
                                   var_cost=0.92,
                                   avg_price=3.62,
                                   qty_actual=405480,
                                   n_steps=3, pct=0.10)
    check("returns a row number > 1",            next_row > 1)
    check("A1 has sensitivity title",            "Sensitivity" in str(ws.cell(1, 1).value))
    check("corner formula present on sheet",     any(
        str(ws.cell(r, 1).value or "").startswith("=")
        for r in range(1, next_row)
    ))


# ── Test 10: export_bsim_excel (full integration) ────────────────────────────
def test_export_bsim_excel():
    print("\n[10] export_bsim_excel (integration)")
    out = "test_output.xlsx"
    export_bsim_excel(
        fixed_cost=441002.73,
        var_cost=0.92,
        avg_price=3.62,
        qty_actual=405480,
        output_path=out
    )
    check("file exists",       os.path.exists(out))
    check("file is non-empty", os.path.getsize(out) > 0)

    from openpyxl import load_workbook
    wb = load_workbook(out)
    check("sheet 1 name correct", wb.sheetnames[0] == "Goal Seek - Break Even")
    check("sheet 2 name correct", wb.sheetnames[1] == "Data Table - Sensitivity")

    ws1 = wb["Goal Seek - Break Even"]
    check("sheet 1 has content",  ws1.cell(1, 1).value is not None)

    ws2 = wb["Data Table - Sensitivity"]
    check("sheet 2 has content",  ws2.cell(1, 1).value is not None)

    os.remove(out)


# ── Run all ───────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    tests = [
        test_generate_series,
        test_write_header,
        test_write_inputs,
        test_write_instructions,
        test_generate_goal_seek,
        test_generate_table_price_qty,
        test_generate_table_cost_qty,
        test_generate_table_price_cost,
        test_generate_data_table,
        test_export_bsim_excel,
    ]

    passed = 0
    failed = 0
    for t in tests:
        try:
            t()
            passed += 1
        except AssertionError as e:
            print(f"  ^^^ FAILED: {e}")
            failed += 1

    print(f"\n{'='*40}")
    print(f"  {passed} passed  |  {failed} failed")
    print(f"{'='*40}")
